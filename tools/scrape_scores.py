import os
import sys
import time
import shutil
import json
from datetime import datetime
from io import StringIO
from typing import List, Dict, Any, Optional, Tuple

import requests
from bs4 import BeautifulSoup
import pandas as pd
from openpyxl import load_workbook

# ----------------------------
# 可配置参数
# ----------------------------
LOGIN_URL = "http://121.229.61.27:8001/xs/cjcx/index.asp"
INPUT_EXCEL = "19班 50人1.xlsx"
SHEET_USERS = "Sheet1"
OUTPUT_EXCEL = "汇总成绩.xlsx"
TIMEOUT = 20
HEADERS = {
    "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/124.0.0.0 Safari/537.36",
    "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,image/avif,image/webp,*/*;q=0.8",
    "Accept-Language": "zh-CN,zh;q=0.9",
    "Connection": "keep-alive",
}

# 若站点登录表单的字段名与默认推断不同，可在此显式指定。
# 例如：{"username": "xm", "password": "mm", "captcha": "yzm"}
FORM_FIELD_NAMES = {
    "username": "adminname",  # 站点表单字段（根据调试文件确定）
    "password": "adminpwd",
    "captcha": None,
}

# 是否需要在控制台手动输入验证码。若站点强制验证码，设为 True
PROMPT_CAPTCHA = False

# 成绩汇总文件路径
SUMMARY_EXCEL = os.path.join("..", "成绩汇总.xlsx")

# 历次成绩目录
EXAMS_DIR = os.path.join("..", "历次成绩")

# 需要提取成绩的学生姓名
TARGET_STUDENT_NAME = "陈泓宇"


def extract_exam_name_from_page(html: str) -> Optional[str]:
    """从登录后的页面中提取考试名称,不依赖正则表达式"""
    soup = BeautifulSoup(html, 'lxml')

    # 方法: 从body开头文本提取
    # 通常考试名称在"返回首页"链接之前
    body = soup.find('body')
    if body:
        # 获取body的直接文本内容(按顺序)
        text_parts = []
        for element in body.descendants:
            if isinstance(element, str):
                text = element.strip()
                if text:
                    text_parts.append(text)
                    # 遇到"返回首页"就停止,考试名称在它前面
                    if '返回首页' in text:
                        break

        # 合并文本并清理
        full_text = ' '.join(text_parts)

        # 移除学校名称等前缀
        prefixes_to_remove = ['无锡市辅仁高级中学', '无锡市', '辅仁高级中学']
        for prefix in prefixes_to_remove:
            if prefix in full_text:
                full_text = full_text.split(prefix)[-1].strip()

        # 移除"返回首页"等后缀
        for suffix in ['返回首页', '首页']:
            if suffix in full_text:
                full_text = full_text.split(suffix)[0].strip()

        # 清理多余空白和特殊符号
        exam_name = ' '.join(full_text.split())
        # 移除可能存在的】、【等符号
        exam_name = exam_name.replace('】', '').replace('【', '').strip()

        if exam_name and len(exam_name) > 5:  # 考试名称至少几个字符
            return exam_name

    return None


def show_menu() -> str:
    """显示交互式菜单并获取用户选择"""
    print("\n" + "="*60)
    print("成绩抓取与汇总系统")
    print("="*60)
    print("请选择操作:")
    print("[1] 完整执行 (抓取成绩 + 汇总 + 生成导入JSON)")
    print("[2] 仅抓取成绩")
    print("[3] 仅成绩汇总")
    print("[4] 仅生成导入JSON (用于网页导入)")
    print("[0] 退出")
    print("="*60)

    while True:
        try:
            choice = input("\n请输入选项 (0-4): ").strip()
            if choice in ['0', '1', '2', '3', '4']:
                return choice
            print("无效选项,请重新输入!")
        except (EOFError, KeyboardInterrupt):
            print("\n\n用户取消操作,退出。")
            sys.exit(0)


def select_exam_file() -> Optional[str]:
    """列出历次成绩目录的xlsx文件供用户选择"""
    xlsx_files = []

    # 扫描历次成绩目录的xlsx文件
    if not os.path.exists(EXAMS_DIR):
        print(f"错误: 历次成绩目录不存在: {EXAMS_DIR}")
        return None

    for filename in os.listdir(EXAMS_DIR):
        if filename.endswith('.xlsx') and not filename.startswith('~$'):
            filepath = os.path.join(EXAMS_DIR, filename)
            if os.path.isfile(filepath):
                xlsx_files.append((filename, filepath))

    if not xlsx_files:
        print("未找到可用的考试文件!")
        return None

    print("\n可用的考试文件:")
    print("="*60)
    for idx, (filename, _) in enumerate(xlsx_files, 1):
        print(f"[{idx}] {filename}")
    print("="*60)

    while True:
        try:
            choice = input(f"\n请选择文件 (1-{len(xlsx_files)}) 或输入 0 取消: ").strip()
            if choice == '0':
                return None

            idx = int(choice) - 1
            if 0 <= idx < len(xlsx_files):
                selected_file = xlsx_files[idx][1]
                selected_name = xlsx_files[idx][0]
                print(f"已选择: {selected_name}")
                return selected_file

            print(f"无效选项,请输入 1-{len(xlsx_files)} 之间的数字!")
        except (ValueError, EOFError, KeyboardInterrupt):
            print("输入无效或用户取消操作!")
            return None


def check_exam_exists(summary_path: str, exam_name: str) -> bool:
    """检查成绩汇总文件中是否已存在该考试"""
    try:
        df = pd.read_excel(summary_path, sheet_name="考试详情")
        return exam_name in df.columns
    except Exception as e:
        print(f"读取成绩汇总文件失败: {e}")
        return False


def ask_overwrite(exam_name: str) -> bool:
    """询问用户是否覆盖已存在的考试成绩"""
    print(f"\n警告: 成绩汇总中已存在考试【{exam_name}】!")
    while True:
        try:
            choice = input("是否覆盖? (y/n): ").strip().lower()
            if choice in ['y', 'yes', '是']:
                return True
            elif choice in ['n', 'no', '否']:
                return False
            print("请输入 y 或 n!")
        except (EOFError, KeyboardInterrupt):
            print("\n用户取消操作。")
            return False


def extract_student_scores(excel_path: str, student_name: str) -> Optional[pd.Series]:
    """从考试文件中提取指定学生的成绩"""
    try:
        df = pd.read_excel(excel_path, sheet_name="汇总")

        # 清理列名: 移除制表符、空格等,并统一字符(其他->其它)
        df.columns = [str(col).strip().replace('\t', '').replace(' ', '').replace('其他', '其它') for col in df.columns]

        # 查找学生
        student_row = df[df['姓名'] == student_name]

        if student_row.empty:
            print(f"警告: 在文件中未找到学生【{student_name}】的成绩!")
            return None

        # 返回清理后的Series
        scores = student_row.iloc[0]
        scores.index = [str(idx).strip().replace('\t', '').replace(' ', '').replace('其他', '其它') for idx in scores.index]
        return scores
    except Exception as e:
        print(f"读取考试文件失败: {e}")
        return None


def update_summary_with_scores(summary_path: str, exam_name: str, student_scores: pd.Series, overwrite: bool = False) -> bool:
    """将学生成绩更新到成绩汇总文件 - 只修改单元格值,保留格式"""
    try:
        # 使用openpyxl打开,保留所有格式
        wb = load_workbook(summary_path)
        ws = wb["考试详情"]

        # 读取第一列(指标名称)建立映射
        indicator_to_row = {}
        for row_idx, row in enumerate(ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=False), start=2):
            indicator_cell = row[0]  # 第一列是指标名称
            if indicator_cell.value:
                indicator = str(indicator_cell.value).strip().replace('\t', '').replace(' ', '').replace('其他', '其它')
                indicator_to_row[indicator] = row_idx

        # 查找或创建考试列
        exam_col = None
        for col_idx, cell in enumerate(ws[1], start=1):  # 第一行是列名
            if cell.value == str(exam_name):
                exam_col = col_idx
                break

        if exam_col is None:
            # 找到最后一列,在其后添加新列
            exam_col = ws.max_column + 1

            # 复制前一列的格式到新列
            from copy import copy
            prev_col = exam_col - 1
            for row_idx in range(1, ws.max_row + 1):
                source_cell = ws.cell(row=row_idx, column=prev_col)
                target_cell = ws.cell(row=row_idx, column=exam_col)

                # 复制格式(不包括值)
                if source_cell.has_style:
                    target_cell.font = copy(source_cell.font)
                    target_cell.border = copy(source_cell.border)
                    target_cell.fill = copy(source_cell.fill)
                    target_cell.number_format = source_cell.number_format
                    target_cell.protection = copy(source_cell.protection)
                    target_cell.alignment = copy(source_cell.alignment)

            # 设置第一行(列名)的值
            ws.cell(row=1, column=exam_col, value=exam_name)
            print(f"  → 创建新列: 第{exam_col}列(已复制格式)")
        else:
            if not overwrite:
                print(f"  → 考试已存在,跳过")
                return False
            print(f"  → 覆盖现有列: 第{exam_col}列")

        # 填充数据
        matched_count = 0
        for indicator, value in student_scores.items():
            indicator_clean = str(indicator).strip().replace('\t', '').replace(' ', '').replace('其他', '其它')

            if indicator_clean in indicator_to_row:
                row_idx = indicator_to_row[indicator_clean]
                ws.cell(row=row_idx, column=exam_col, value=value)
                matched_count += 1

        print(f"  → 匹配了 {matched_count} 个指标")

        # 保存文件
        wb.save(summary_path)

        print(f"✅ 已成功更新成绩汇总: {exam_name}")
        return True

    except Exception as e:
        print(f"更新成绩汇总失败: {e}")
        import traceback
        traceback.print_exc()
        return False


def scrape_scores_with_exam_name() -> Tuple[bool, Optional[str]]:
    """执行抓取流程,返回(是否成功, 考试名称)"""
    output_excel = OUTPUT_EXCEL
    users_df = read_users_from_excel(INPUT_EXCEL, SHEET_USERS)
    template_df: Optional[pd.DataFrame] = None
    all_rows: List[pd.DataFrame] = []
    debug_dir = None
    exam_name = None
    count = 0

    for idx, row in users_df.iterrows():
        username_value = None
        candidate_cols = ["姓名", "学号", "账号", "考籍号", "准考证号", "用户名"]
        for col in candidate_cols:
            if col in row.index and str(row[col]).strip() not in ("", "nan", "None"):
                username_value = str(row[col]).strip()
                break

        pwd = str(row.get("密码", "")).strip()
        name = str(row.get("姓名", username_value or "")).strip()

        if not name or not pwd:
            continue

        print(f"[{idx+1}/{len(users_df)}] 登录并抓取：{name}")

        with requests.Session() as sess:
            try:
                df = scrape_for_user(sess, username_value or name, pwd, debug_dir=debug_dir, user_idx=idx+1)

                # 首次成功登录后提取考试名称
                if df is not None and exam_name is None:
                    # 重新获取页面来提取考试名称
                    try:
                        r = sess.get(LOGIN_URL, headers=HEADERS, timeout=TIMEOUT)
                        _fix_response_encoding(r)
                        payload, action, method = build_login_payload(r.text, username_value or name, pwd, None)
                        submit_url = resolve_url(LOGIN_URL, action) if action else LOGIN_URL

                        from urllib.parse import urlencode
                        encoded_form = urlencode(payload, encoding="gb18030", doseq=True)
                        headers = dict(HEADERS)
                        headers["Referer"] = LOGIN_URL
                        headers["Content-Type"] = "application/x-www-form-urlencoded; charset=gb18030"

                        r2 = sess.post(submit_url, data=encoded_form, headers=headers, timeout=TIMEOUT, allow_redirects=True)
                        _fix_response_encoding(r2)

                        exam_name = extract_exam_name_from_page(r2.text)
                        if exam_name:
                            print(f"  → 检测到考试名称: {exam_name}")
                    except Exception as e:
                        print(f"  → 提取考试名称失败: {e}")

                if df is None or df.empty:
                    print(f"  - 未获取到表格：{name}")
                    continue

                # 首位同学保留最后两行(包含表头信息),其余保留最后一行
                if count == 0 and len(df) >= 2:
                    df = df.tail(2).reset_index(drop=True)
                else:
                    df = pick_score_row(df, mode="last")

                df = coerce_numeric_like(df, exclude_cols=["姓名", "学号", "账号", "准考证号", "考籍号", "用户名"])

                # 如果表格中已经有"姓名"列,直接更新;否则插入
                if "姓名" in df.columns:
                    df.loc[:, "姓名"] = name
                else:
                    df.insert(0, "姓名", name)

                aligned = ensure_output_columns(template_df, df)
                if "姓名" not in aligned.columns:
                    aligned.insert(0, "姓名", name)
                else:
                    aligned.loc[:, "姓名"] = name
                df = aligned

                all_rows.append(df)
                time.sleep(0.8)

                if debug_dir is not None:
                    try:
                        df.to_csv(os.path.join(debug_dir, f"{idx+1:03d}_selected_table.csv"), index=False, encoding="utf-8-sig")
                    except Exception:
                        pass

            except Exception as e:
                print(f"  - 失败：{name}，{e}")
                continue

        count += 1

    if not all_rows:
        print("未抓取到任何数据，退出。")
        return False, None

    result = pd.concat(all_rows, ignore_index=True)

    # 写出汇总成绩.xlsx
    with pd.ExcelWriter(output_excel, engine="openpyxl") as writer:
        result.to_excel(writer, index=False, sheet_name="汇总")
    print(f"✅ 已写出：{output_excel}，共 {len(result)} 行。")

    return True, exam_name


def move_exam_file(exam_name: str) -> Optional[str]:
    """重命名并移动考试文件到历次成绩目录"""
    if not exam_name:
        print("⚠ 考试名称为空,无法重命名文件!")
        return None

    source_file = OUTPUT_EXCEL
    target_filename = f"{exam_name}.xlsx"
    target_path = os.path.join(EXAMS_DIR, target_filename)

    # 检查目标文件是否已存在
    if os.path.exists(target_path):
        print(f"⚠ 目标文件已存在: {target_filename}")
        while True:
            try:
                choice = input("是否覆盖? (y/n): ").strip().lower()
                if choice in ['y', 'yes', '是']:
                    break
                elif choice in ['n', 'no', '否']:
                    print("取消文件移动。")
                    return None
                print("请输入 y 或 n!")
            except (EOFError, KeyboardInterrupt):
                print("\n用户取消操作。")
                return None

    try:
        shutil.move(source_file, target_path)
        print(f"✅ 文件已移动到: {target_path}")
        return target_path
    except Exception as e:
        print(f"❌ 移动文件失败: {e}")
        return None


def update_summary(exam_file_path: str, exam_name: str) -> bool:
    """从考试文件更新成绩汇总"""
    print(f"\n开始从 {exam_file_path} 提取【{TARGET_STUDENT_NAME}】的成绩...")

    # 提取学生成绩
    student_scores = extract_student_scores(exam_file_path, TARGET_STUDENT_NAME)
    if student_scores is None:
        return False

    print(f"✅ 成功提取学生成绩")

    # 检查是否已存在
    if check_exam_exists(SUMMARY_EXCEL, exam_name):
        if not ask_overwrite(exam_name):
            print("❌ 用户取消操作,未更新成绩汇总。")
            return False
        overwrite = True
    else:
        overwrite = False

    # 更新成绩汇总
    return update_summary_with_scores(SUMMARY_EXCEL, exam_name, student_scores, overwrite)


def export_to_json(excel_path: str, exam_name: str) -> Optional[str]:
    """将考试 Excel 转换为 Web 系统所需的 JSON 格式"""
    print(f"\n正在生成导入 JSON...")
    try:
        # 读取 Excel
        df = pd.read_excel(excel_path, sheet_name="汇总")
        
        # 清理列名
        df.columns = [str(col).strip() for col in df.columns]
        
        # 识别科目 (过滤掉姓名、账号、总分等非科目列)
        EXCLUDE = ['姓名', '账号', '学号', '考籍号', '准考证号', '用户名', '密码', '总分', '名次', '班级', '年级排名', '班级排名']
        SUBJECTS = [col for col in df.columns if col not in EXCLUDE and '排名' not in col and '均分' not in col]
        
        import_data = {
            "examName": exam_name,
            "examDate": datetime.now().strftime("%Y-%m-%d"),
            "data": []
        }

        for _, row in df.iterrows():
            # 基础信息
            student = {
                "student_name": str(row.get('姓名', '')),
                "class_name": str(row.get('班级', '未知')),
                "total_score": float(row.get('总分', 0)),
                "grade_rank": int(row.get('年级排名', 0)) if pd.notnull(row.get('年级排名')) else 0,
                "class_rank": int(row.get('班级排名', 0)) if pd.notnull(row.get('班级排名')) else 0,
                "subjects": []
            }

            # 提取各科
            for sub in SUBJECTS:
                student["subjects"].append({
                    "subject": sub,
                    "score": float(row.get(sub, 0)),
                    "grade_rank": int(row.get(f'{sub}年级排名', 0)) if f'{sub}年级排名' in df.columns else 0,
                    "class_rank": int(row.get(f'{sub}班级排名', 0)) if f'{sub}班级排名' in df.columns else 0,
                    "class_avg": float(row.get(f'{sub}班级均分', 0)) if f'{sub}班级均分' in df.columns else 0
                })
            
            import_data["data"].append(student)

        # 保存
        output_path = os.path.join(os.path.dirname(excel_path), f"{exam_name}_import.json")
        with open(output_path, 'w', encoding='utf-8') as f:
            json.dump(import_data, f, ensure_ascii=False, indent=2)
        
        print(f"✅ JSON 生成成功: {os.path.basename(output_path)}")
        return output_path
    except Exception as e:
        print(f"❌ 生成 JSON 失败: {e}")
        return None


def _fix_response_encoding(resp: requests.Response) -> None:
    """确保中文站点编码被正确解码，优先使用apparent_encoding，其次回退gb18030。"""
    try:
        if not resp.encoding or resp.encoding.lower() in ("iso-8859-1", "ascii"):
            enc = getattr(resp, "apparent_encoding", None)
            if enc:
                resp.encoding = enc
            else:
                resp.encoding = "gb18030"
    except Exception:
        # 任意异常回退统一编码
        resp.encoding = "gb18030"


def read_users_from_excel(path: str, sheet_name: str) -> pd.DataFrame:
    if not os.path.exists(path):
        raise FileNotFoundError(f"未找到Excel文件: {path}")
    df = pd.read_excel(path, sheet_name=sheet_name)
    # 尝试识别列名（允许大小写或空白差异）
    name_col = None
    pwd_col = None
    for c in df.columns:
        lc = str(c).strip().lower()
        if lc in ("姓名", "name", "用户名", "user", "账号", "学号"):
            name_col = c if name_col is None else name_col
        if lc in ("密码", "pass", "password", "pwd"):
            pwd_col = c if pwd_col is None else pwd_col
    if name_col is None or pwd_col is None:
        # 退而求其次，取前两列
        if len(df.columns) >= 2:
            name_col, pwd_col = df.columns[0], df.columns[1]
        else:
            raise ValueError("无法识别sheet1中的姓名与密码列，请确认表头或将其置于前两列")
    # 仅保留两列并重命名
    out = df[[name_col, pwd_col]].copy()
    out.columns = ["姓名", "密码"]
    out = out.dropna(how="any")
    return out


def detect_form_fields(html: str) -> Dict[str, Optional[str]]:
    soup = BeautifulSoup(html, "lxml")
    form = soup.find("form")
    if not form:
        return {"username": None, "password": None, "captcha": None, "action": None, "method": "post"}
    inputs = form.find_all("input")
    username_field = None
    password_field = None
    captcha_field = None
    # 优先根据常用命名匹配
    for inp in inputs:
        name_attr = (inp.get("name") or "").lower()
        itype = (inp.get("type") or "").lower()
        placeholder = (inp.get("placeholder") or "").lower()
        if not username_field and ("user" in name_attr or name_attr in ("xm", "xh", "name", "zh", "account", "xj", "kzh") or "account" in name_attr or (itype == "text" and ("user" in placeholder or "学号" in placeholder or "账号" in placeholder))):
            username_field = inp.get("name")
        if not password_field and (itype == "password" or "pass" in name_attr or name_attr in ("mm", "pwd")):
            password_field = inp.get("name")
        if not captcha_field and ("yzm" in name_attr or "captcha" in name_attr or "verify" in name_attr or name_attr == "code"):
            captcha_field = inp.get("name")
    # 若仍未找到，按出现顺序：第一个 text 当用户名，第二个 text 或 password 当密码，第三个可能为验证码
    text_like = [i for i in inputs if (i.get("type") or "").lower() in ("text", "password", "")]
    if not username_field and text_like:
        username_field = text_like[0].get("name")
    if not password_field and len(text_like) >= 2:
        password_field = text_like[1].get("name")
    if not captcha_field and len(text_like) >= 3:
        captcha_field = text_like[2].get("name")

    action = form.get("action")
    method = (form.get("method") or "post").lower()
    return {
        "username": username_field,
        "password": password_field,
        "captcha": captcha_field,
        "action": action,
        "method": method,
    }


def build_login_payload(html: str, username: str, password: str, captcha: Optional[str]) -> (Dict[str, Any], str, str):
    detected = detect_form_fields(html)
    # 允许通过FORM_FIELD_NAMES覆盖
    username_field = FORM_FIELD_NAMES.get("username") or detected.get("username")
    password_field = FORM_FIELD_NAMES.get("password") or detected.get("password")
    captcha_field = FORM_FIELD_NAMES.get("captcha") or detected.get("captcha")
    if not username_field or not password_field:
        raise ValueError("无法从登录页推断用户名/密码字段名，请在脚本顶部显式配置FORM_FIELD_NAMES")

    soup = BeautifulSoup(html, "lxml")
    form = soup.find("form")
    action = detected.get("action")
    method = detected.get("method") or "post"
    # 收集所有隐藏字段
    payload: Dict[str, Any] = {}
    if form:
        for inp in form.find_all("input"):
            name_attr = inp.get("name")
            if not name_attr:
                continue
            itype = (inp.get("type") or "").lower()
            val = inp.get("value")
            if itype == "hidden":
                payload[name_attr] = val if val is not None else ""
    # 填入用户字段
    payload[username_field] = username
    payload[password_field] = password
    if captcha_field and captcha is not None:
        payload[captcha_field] = captcha

    return payload, action or "", method
def is_login_success(html: str) -> bool:
    """粗略判定是否已登录：页面不再出现密码输入框/登录按钮，且包含成绩相关关键词。"""
    soup = BeautifulSoup(html, "lxml")
    # 若仍有 password 输入，基本可判定未登录
    if soup.find("input", {"type": "password"}):
        return False
    text = soup.get_text(" ", strip=True)
    for k in ("成绩", "分数", "课程", "科目", "总分", "试卷"):
        if k in text:
            return True
    # 没有明显成绩关键词，也可能登录后还需点击进入；放宽处理
    return True



def _parse_table_consistent(table_tag) -> Optional[pd.DataFrame]:
    """解析HTML <table> 为等宽DataFrame，处理多行表头情况"""
    try:
        # 先尝试用pandas读取,使用header=0
        tables = pd.read_html(StringIO(str(table_tag)), header=0)
        if tables and not tables[0].empty:
            df = tables[0]
            # 检查第一行是否全为数字(如果是,说明第一行是序号,第二行才是表头)
            if len(df) > 0:
                first_row = df.iloc[0]
                # 如果第一行的所有值都是数字或NaN,说明是序号行
                is_number_row = all(
                    pd.isna(val) or str(val).isdigit() or (isinstance(val, (int, float)) and val == int(val))
                    for val in first_row.values
                )
                if is_number_row:
                    # 跳过第一行,重新读取
                    tables = pd.read_html(StringIO(str(table_tag)), header=[0, 1])
                    if tables and not tables[0].empty:
                        return tables[0]
            return df
    except Exception:
        pass

    # 如果pandas读取失败,尝试手动解析
    rows: List[List[str]] = []
    max_cols = 0
    for tr in table_tag.find_all("tr"):
        row: List[str] = []
        for cell in tr.find_all(["td", "th"]):
            text = cell.get_text(strip=True)
            colspan = int(cell.get("colspan") or 1)
            for _ in range(colspan):
                row.append(text)
        max_cols = max(max_cols, len(row))
        rows.append(row)

    if not rows:
        return None

    # 对齐每行列数
    rows = [r + [""] * (max_cols - len(r)) for r in rows]

    # 检查第一行是否全是数字(序号行)
    if len(rows) >= 3:
        first_row = rows[0]
        is_number_row = all(cell == "" or cell.isdigit() for cell in first_row)
        if is_number_row:
            # 第二行才是真正的表头
            if any("\u4e00" <= ch <= "\u9fff" for cell in rows[1] for ch in cell):
                df = pd.DataFrame(rows[2:], columns=rows[1])
                return df

    # 尝试用第一行作为表头
    if len(rows) >= 2:
        first_row = rows[0]
        chinese_count = sum(1 for cell in first_row if any("\u4e00" <= ch <= "\u9fff" for ch in cell))
        if chinese_count >= 3:
            df = pd.DataFrame(rows[1:], columns=first_row)
            return df

    # 使用默认列名
    df = pd.DataFrame(rows)
    return df


def extract_first_table(html: str) -> Optional[pd.DataFrame]:
    soup = BeautifulSoup(html, "lxml")
    table = soup.find("table")
    if not table:
        return None
    df = _parse_table_consistent(table)
    return df


def resolve_url(base_url: str, action: str) -> str:
    from urllib.parse import urljoin
    return urljoin(base_url, action)


KEYWORDS = ["成绩", "分数", "分", "科目", "课程", "总分", "平均分", "名次", "班级", "学号"]


def select_best_table(html: str) -> Optional[pd.DataFrame]:
    """在页面中选择最可能是成绩的表格：优先选择class=a2的内层表格"""
    soup = BeautifulSoup(html, "lxml")

    # 优先查找class="a2"的表格(内层成绩表格)
    a2_table = soup.find('table', class_='a2')
    if a2_table:
        try:
            dfs = pd.read_html(StringIO(str(a2_table)), header=0)
            if dfs and not dfs[0].empty and len(dfs[0].columns) > 20:
                return dfs[0]
        except Exception:
            pass

    # 如果找不到a2表格,使用原有逻辑
    tables = soup.find_all("table")
    if not tables:
        return None
    best_score = -1
    best_df: Optional[pd.DataFrame] = None
    for t in tables:
        df: Optional[pd.DataFrame] = None
        try:
            tmp = pd.read_html(StringIO(str(t)), header=0)
            if tmp:
                df = tmp[0]
        except Exception:
            df = None
        if df is None or df.empty:
            # 简易兜底
            try:
                headers: List[str] = []
                rows: List[List[str]] = []
                thead = t.find("thead")
                if thead:
                    headers = [th.get_text(strip=True) for th in thead.find_all("th")]
                tbody = t.find("tbody") or t
                for tr in tbody.find_all("tr"):
                    cols = tr.find_all(["td", "th"])
                    if not headers:
                        headers = [f"列{i+1}" for i in range(len(cols))]
                    rows.append([c.get_text(strip=True) for c in cols])
                if rows:
                    df = pd.DataFrame(rows, columns=headers)
            except Exception:
                df = None
        if df is None or df.empty:
            continue
        # 评分：优先选择列数多的表格(内层成绩表格列数通常>50)
        rows_n, cols_n = df.shape
        score = cols_n * 100 + rows_n
        # 关键词匹配加分
        text_blob = " ".join([" ".join(map(str, df.columns.tolist()))] + [" ".join(map(str, r)) for r in df.head(3).values.tolist()])
        kw_hits = sum(1 for k in KEYWORDS if k in text_blob)
        score += kw_hits * 10
        # 列数太少，降权
        if cols_n <= 5:
            score -= 1000
        if score > best_score:
            best_score = score
            best_df = df
    return best_df


def pick_first_df(*candidates: Optional[pd.DataFrame]) -> Optional[pd.DataFrame]:
    for d in candidates:
        try:
            if d is not None and not d.empty:
                return d
        except Exception:
            continue
    return None


def ensure_output_columns(_: Optional[pd.DataFrame], data_df: pd.DataFrame) -> pd.DataFrame:
    """不再参考Sheet2：仅将“姓名”列放到首列（如存在）。"""
    cols = list(data_df.columns)
    if "姓名" in cols:
        cols.remove("姓名")
        cols = ["姓名"] + cols
    return data_df.reindex(columns=cols)


def pick_score_row(data_df: pd.DataFrame, mode: str = "third") -> pd.DataFrame:
    """根据模式挑选成绩行：
    - third: 取第3行（0基索引为2），若不足3行则取最后一行
    - last:  取最后一行
    - none:  不筛选，返回原表
    返回仅包含1行的数据框。
    """
    if data_df is None or data_df.empty:
        return data_df
    if mode == "none":
        return data_df
    if mode == "last":
        return data_df.tail(1).reset_index(drop=True)
    # default: third
    if len(data_df) >= 3:
        return data_df.iloc[[2]].reset_index(drop=True)
    return data_df.tail(1).reset_index(drop=True)

def coerce_numeric_like(df: pd.DataFrame, *, exclude_cols: Optional[List[str]] = None, threshold: float = 0.6) -> pd.DataFrame:
    """将看起来是数字的列转换为数值类型。
    - exclude_cols: 不尝试转换的列名列表（例如 姓名、学号等）
    - threshold: 可成功转为数值的比例阈值，达到则整列转换，否则保持原样
    """
    if df is None or df.empty:
        return df
    exclude_cols = exclude_cols or []
    out = df.copy()
    for col in out.columns:
        if col in exclude_cols:
            continue
        series = out[col].astype(str)
        # 清洗：去掉千分位逗号、空格、中文空格、尾部的“分”字等
        cleaned = (
            series.str.replace("\u3000", "", regex=False)
                  .str.replace(",", "", regex=False)
                  .str.replace(" ", "", regex=False)
                  .str.replace("分", "", regex=False)
                  .str.replace("—", "", regex=False)
                  .str.replace("-", "-", regex=False)
        )
        numeric = pd.to_numeric(cleaned, errors="coerce")
        ratio = numeric.notna().mean()
        count_numeric = int(numeric.notna().sum())
        n_rows = len(numeric)
        # 行数很少（<=2）且至少有1个可转，则转换
        if n_rows <= 2 and count_numeric >= 1:
            out[col] = numeric
            continue
        # 常规：多数可转换时整体转换
        if ratio >= threshold:
            out[col] = numeric
    return out

def scrape_for_user(session: requests.Session, username: str, password: str, *, debug_dir: Optional[str] = None, user_idx: Optional[int] = None) -> Optional[pd.DataFrame]:
    r = session.get(LOGIN_URL, headers=HEADERS, timeout=TIMEOUT)
    r.raise_for_status()
    _fix_response_encoding(r)
    if debug_dir is not None and user_idx is not None:
        try:
            with open(os.path.join(debug_dir, f"{user_idx:03d}_login_page.html"), "w", encoding=r.encoding or "utf-8") as f:
                f.write(r.text)
        except Exception:
            pass
    captcha_value: Optional[str] = None
    if PROMPT_CAPTCHA:
        # 简易提示：若页面包含“验证码”字样，则提示输入
        if "验证码" in r.text or "驗證碼" in r.text or "captcha" in r.text.lower():
            print(f"检测到可能的验证码，账号 {username} 需要手动输入（若无请输入直接回车）：")
            try:
                captcha_value = input("请输入验证码: ").strip()
            except EOFError:
                captcha_value = None
    payload, action, method = build_login_payload(r.text, username, password, captcha_value)
    submit_url = resolve_url(LOGIN_URL, action) if action else LOGIN_URL
    # 补充常见 Referer
    headers = dict(HEADERS)
    headers["Referer"] = LOGIN_URL
    # 使用GB2312对表单进行编码，避免中文姓名在目标站点被误解码
    from urllib.parse import urlencode
    encoded_form = urlencode(payload, encoding="gb18030", doseq=True)
    headers["Content-Type"] = "application/x-www-form-urlencoded; charset=gb18030"
    if method == "post":
        r2 = session.post(submit_url, data=encoded_form, headers=headers, timeout=TIMEOUT, allow_redirects=True)
    else:
        r2 = session.get(submit_url + ("?" + encoded_form if encoded_form else ""), headers=headers, timeout=TIMEOUT, allow_redirects=True)
    r2.raise_for_status()
    _fix_response_encoding(r2)
    if debug_dir is not None and user_idx is not None:
        try:
            with open(os.path.join(debug_dir, f"{user_idx:03d}_after_login.html"), "w", encoding=r2.encoding or "utf-8") as f:
                f.write(r2.text)
        except Exception:
            pass

    # 登录成败快速判定
    if not is_login_success(r2.text):
        if debug_dir is not None and user_idx is not None:
            try:
                with open(os.path.join(debug_dir, f"{user_idx:03d}_login_failed.txt"), "w", encoding="utf-8") as f:
                    f.write("登录疑似失败：页面仍含密码输入或未出现成绩关键词。\n")
                    f.write(f"POST URL: {submit_url}\n")
                    f.write(f"Status: {r2.status_code}\n")
                    f.write(f"Final URL: {r2.url}\n")
            except Exception:
                pass
        return None

    # 登录后页面直接有表格，或需跳转；先尝试当前页
    df = pick_first_df(select_best_table(r2.text), extract_first_table(r2.text))
    if df is not None and not df.empty:
        return df
    # 若有meta refresh或a链接提示成绩页，尝试跟随
    soup = BeautifulSoup(r2.text, "lxml")
    link = soup.find("a")
    if link and link.get("href"):
        url = resolve_url(submit_url, link.get("href"))
        r3 = session.get(url, headers=HEADERS, timeout=TIMEOUT)
        if r3.ok:
            _fix_response_encoding(r3)
            df = pick_first_df(select_best_table(r3.text), extract_first_table(r3.text))
            if df is not None and not df.empty:
                return df
    # 最后兜底
    return pick_first_df(select_best_table(r2.text), extract_first_table(r2.text))


def main() -> None:
    """主函数 - 交互式菜单"""
    # 显示菜单并获取用户选择
    choice = show_menu()

    if choice == '0':
        print("退出程序。")
        return

    elif choice == '2':
        # 仅抓取成绩
        print("\n开始抓取成绩...")
        success, exam_name = scrape_scores_with_exam_name()

        if success and exam_name:
            # 移动文件到上级目录
            moved_file = move_exam_file(exam_name)
            if moved_file:
                print(f"\n✅ 抓取完成! 文件已保存为: {moved_file}")
            else:
                print(f"\n⚠ 抓取完成,但文件移动失败。文件仍在: {OUTPUT_EXCEL}")
        elif success:
            print(f"\n⚠ 抓取完成,但未能提取考试名称。文件保存在: {OUTPUT_EXCEL}")
        else:
            print("\n❌ 抓取失败!")

    elif choice == '3':
        # 仅成绩汇总
        exam_file = select_exam_file()
        if not exam_file:
            print("未选择文件,退出。")
            return

        # 从文件名提取考试名称
        exam_name = os.path.basename(exam_file).replace('.xlsx', '')
        print(f"使用考试名称: {exam_name}")

        # 更新成绩汇总
        if update_summary(exam_file, exam_name):
            print(f"\n✅ 成绩汇总更新完成!")
        else:
            print("\n❌ 成绩汇总更新失败!")

    elif choice == '1':
        # 完整执行
        print("\n开始完整执行流程...")

        # 步骤1: 抓取成绩
        print("\n[步骤 1/3] 抓取成绩...")
        success, exam_name = scrape_scores_with_exam_name()

        if not success:
            print("\n❌ 抓取失败,终止流程。")
            return

        if not exam_name:
            print(f"\n⚠ 未能提取考试名称,使用默认名称。文件保存在: {OUTPUT_EXCEL}")
            return

        # 步骤2: 移动文件
        print(f"\n[步骤 2/3] 移动文件...")
        moved_file = move_exam_file(exam_name)
        if not moved_file:
            print("\n❌ 文件移动失败,终止流程。")
            return

        # 步骤3: 更新成绩汇总
        print(f"\n[步骤 3/4] 更新成绩汇总...")
        update_summary(moved_file, exam_name)

        # 步骤4: 生成 JSON
        print(f"\n[步骤 4/4] 生成导入 JSON...")
        json_file = export_to_json(moved_file, exam_name)

        print(f"\n{'='*60}")
        print("✅ 完整执行成功!")
        print(f"考试文件: {moved_file}")
        if json_file:
            print(f"导入 JSON: {json_file} (请在网页端导入)")
        print(f"已更新【{TARGET_STUDENT_NAME}】的成绩到成绩汇总.xlsx")
        print(f"{'='*60}")

    elif choice == '4':
        # 仅生成 JSON
        exam_file = select_exam_file()
        if not exam_file:
            print("未选择文件,退出。")
            return
        
        exam_name = os.path.basename(exam_file).replace('.xlsx', '')
        export_to_json(exam_file, exam_name)


# 保留原有的main函数作为兼容性(已废弃)
# 如需使用旧版本无菜单模式,请手动调用 main_old()


if __name__ == "__main__":
    main()


